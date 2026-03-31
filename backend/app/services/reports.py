"""
Report generation service for PDF and Excel exports.
"""

from datetime import datetime, date
from typing import List, Dict, Any, Optional
from io import BytesIO
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage

from ..db.supabase import db
from ..models import BusinessCategory, SupplierStatus
from ..core.timezone import get_cat_now, format_cat_datetime


class ReportService:
    """Service for generating supplier reports in various formats."""
    
    def __init__(self):
        self.company_name = "Rainbow Tourism Group"
        self.report_title = "Supplier Report"
    
    async def get_filtered_suppliers(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[List[SupplierStatus]] = None,
        category: Optional[List[BusinessCategory]] = None,
        location: Optional[str] = None,
        min_years: Optional[int] = None,
        max_years: Optional[int] = None,
        sort_by: str = "company_name",
        sort_order: str = "asc",
    ) -> List[Dict[str, Any]]:
        """
        Get filtered list of suppliers based on criteria.
        """
        # Fetch only the columns actually consumed by the report renderer.
        # Always scoped to APPROVED — the DB filter avoids transferring rows
        # that are discarded unconditionally in every report type.
        _SUPPLIER_COLUMNS = (
            "id,status,company_name,business_category,"
            "city,country,state_province,street_address,postal_code,"
            "years_in_business,is_small_scale_farmer,business_size,"
            "submitted_at,created_at,updated_at,"
            "contact_person_name,contact_person_title,email,phone,"
            "registration_number,tax_id,website"
        )
        result = (
            db.client.table("suppliers")
            .select(_SUPPLIER_COLUMNS)
            .eq("status", "APPROVED")
            .execute()
        )
        suppliers = result.data if result.data else []

        # Enrich each supplier with all their categories from supplier_categories
        if suppliers:
            all_ids = [s["id"] for s in suppliers]
            cats_resp = db.client.table("supplier_categories").select("supplier_id, category").in_("supplier_id", all_ids).execute()
            cats_by_supplier: Dict[str, List[str]] = {}
            for row in (cats_resp.data or []):
                cats_by_supplier.setdefault(row["supplier_id"], []).append(row["category"])
            for s in suppliers:
                s["business_categories"] = cats_by_supplier.get(s["id"]) or [s["business_category"]]

        # Apply remaining filters (date range, category, location, years)
        filtered = []
        for supplier in suppliers:
            # Date filter (submitted_at with fallback to created_at)
            if start_date or end_date:
                supplier_date_str = supplier.get("submitted_at") or supplier.get("created_at")
                if supplier_date_str:
                    try:
                        supplier_date = datetime.fromisoformat(supplier_date_str.replace('Z', '+00:00')).date()
                        if start_date and supplier_date < start_date:
                            continue
                        if end_date and supplier_date > end_date:
                            continue
                    except Exception:
                        pass
            
            # Category filter — check all the supplier's categories, not just the primary one
            if category:
                filter_values = [c.value for c in category]
                supplier_cats = supplier.get("business_categories") or [supplier.get("business_category")]
                if not any(c in filter_values for c in supplier_cats):
                    continue
            
            # Location filter (check both city and country)
            if location:
                location_lower = location.lower()
                city = (supplier.get("city") or "").lower()
                country = (supplier.get("country") or "").lower()
                if location_lower not in city and location_lower not in country:
                    continue
            
            # Years in business filter
            years = supplier.get("years_in_business")
            if years is not None:
                if min_years is not None and years < min_years:
                    continue
                if max_years is not None and years > max_years:
                    continue
            
            filtered.append(supplier)

        # ── Sorting ────────────────────────────────────────────────────────────
        reverse = sort_order.lower() == "desc"

        if sort_by == "category_then_company":
            # Primary: business_category A→Z; Secondary: company_name A→Z
            filtered.sort(
                key=lambda s: (
                    (s.get("business_category") or "").lower(),
                    (s.get("company_name") or "").lower(),
                )
            )
        elif sort_by == "company_name":
            filtered.sort(key=lambda s: (s.get("company_name") or "").lower(), reverse=reverse)
        elif sort_by == "business_category":
            filtered.sort(key=lambda s: (s.get("business_category") or "").lower(), reverse=reverse)
        elif sort_by == "status":
            filtered.sort(key=lambda s: (s.get("status") or "").lower(), reverse=reverse)
        elif sort_by in ("submitted_at", "created_at"):
            filtered.sort(
                key=lambda s: (s.get(sort_by) or ""),
                reverse=reverse,
            )
        # else: keep original DB order

        return filtered
    
    # ──────────────────────────────────────────────────────────────────────
    # Sustainability helper
    # ──────────────────────────────────────────────────────────────────────

    async def _build_enhanced_sustainability_metrics(
        self,
        suppliers: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """Build enhanced sustainability metrics aligned to the dashboard logic."""
        total_suppliers = len(suppliers)
        if total_suppliers == 0:
            return {
                "total_suppliers": 0,
                "women_owned_count": 0,
                "women_owned_pct": 0.0,
                "youth_owned_count": 0,
                "youth_owned_pct": 0.0,
                "small_scale_farmer_count": 0,
                "small_scale_farmer_pct": 0.0,
                "female_directors": 0,
                "female_director_pct": 0.0,
                "youth_directors": 0,
                "youth_director_pct": 0.0,
                "total_directors": 0,
                "small_enterprises": 0,
                "medium_enterprises": 0,
                "large_enterprises": 0,
                "unknown_size_count": 0,
                "sme_total": 0,
                "sme_pct": 0.0,
                "business_size_distribution": [],
            }

        supplier_ids = [s.get("id") for s in suppliers if s.get("id")]
        esg_rows: List[Dict[str, Any]] = []
        if supplier_ids:
            esg_result = (
                db.client.table("vw_esg_supplier_summary")
                .select(
                    "id,business_size,is_small_scale_farmer,female_director_count,youth_director_count,key_person_count"
                )
                .in_("id", supplier_ids)
                .execute()
            )
            esg_rows = esg_result.data or []

        esg_by_id = {row["id"]: row for row in esg_rows if row.get("id")}

        women_owned_count = 0
        youth_owned_count = 0
        small_scale_farmer_count = 0
        female_directors = 0
        youth_directors = 0
        total_directors = 0

        size_counts: Dict[str, int] = {
            "SMALL": 0,
            "MEDIUM": 0,
            "LARGE": 0,
            "UNKNOWN": 0,
        }

        for supplier in suppliers:
            supplier_id = supplier.get("id")
            esg_row = esg_by_id.get(supplier_id, {})

            female_count = esg_row.get("female_director_count") or 0
            youth_count = esg_row.get("youth_director_count") or 0
            key_person_count = esg_row.get("key_person_count") or 0

            if female_count > 0:
                women_owned_count += 1
            if youth_count > 0:
                youth_owned_count += 1

            is_farmer = esg_row.get("is_small_scale_farmer")
            if is_farmer is None:
                is_farmer = supplier.get("is_small_scale_farmer")
            if bool(is_farmer):
                small_scale_farmer_count += 1

            female_directors += female_count
            youth_directors += youth_count
            total_directors += key_person_count

            raw_size = esg_row.get("business_size") or supplier.get("business_size") or "UNKNOWN"
            size_value = str(raw_size).upper()
            if size_value not in size_counts:
                size_value = "UNKNOWN"
            size_counts[size_value] += 1

        def _safe_pct(part: int, whole: int) -> float:
            return round((part / whole) * 100, 2) if whole > 0 else 0.0

        small_enterprises = size_counts["SMALL"]
        medium_enterprises = size_counts["MEDIUM"]
        large_enterprises = size_counts["LARGE"]
        unknown_size_count = size_counts["UNKNOWN"]
        sme_total = small_enterprises + medium_enterprises

        business_size_distribution = [
            {
                "business_size": "SMALL",
                "supplier_count": small_enterprises,
                "pct": _safe_pct(small_enterprises, total_suppliers),
            },
            {
                "business_size": "MEDIUM",
                "supplier_count": medium_enterprises,
                "pct": _safe_pct(medium_enterprises, total_suppliers),
            },
            {
                "business_size": "LARGE",
                "supplier_count": large_enterprises,
                "pct": _safe_pct(large_enterprises, total_suppliers),
            },
            {
                "business_size": "UNKNOWN",
                "supplier_count": unknown_size_count,
                "pct": _safe_pct(unknown_size_count, total_suppliers),
            },
        ]

        return {
            "total_suppliers": total_suppliers,
            "women_owned_count": women_owned_count,
            "women_owned_pct": _safe_pct(women_owned_count, total_suppliers),
            "youth_owned_count": youth_owned_count,
            "youth_owned_pct": _safe_pct(youth_owned_count, total_suppliers),
            "small_scale_farmer_count": small_scale_farmer_count,
            "small_scale_farmer_pct": _safe_pct(small_scale_farmer_count, total_suppliers),
            "female_directors": female_directors,
            "female_director_pct": _safe_pct(female_directors, total_directors),
            "youth_directors": youth_directors,
            "youth_director_pct": _safe_pct(youth_directors, total_directors),
            "total_directors": total_directors,
            "small_enterprises": small_enterprises,
            "medium_enterprises": medium_enterprises,
            "large_enterprises": large_enterprises,
            "unknown_size_count": unknown_size_count,
            "sme_total": sme_total,
            "sme_pct": _safe_pct(sme_total, total_suppliers),
            "business_size_distribution": business_size_distribution,
        }

    def _build_sustainability_by_category(
        self,
        suppliers: List[Dict[str, Any]],
        sustainability_docs: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        Build per-category sustainability stats.

        Returns:
          {
            "total_suppliers": int,
            "total_submitters": int,
            "overall_rate": float,
            "by_category": [
              {
                "category": str,          # raw DB value, e.g. "FOOD_BEVERAGE"
                "display": str,           # human label, e.g. "Food & Beverage"
                "total": int,
                "submitters": int,
                "rate": float,
                "docs": [ { supplier, document_type, verification_status, uploaded_at }, ... ]
              },
              ...  (sorted by rate desc, then display name asc)
            ],
          }
        """
        # Count total suppliers per category (multi-category aware)
        cat_totals: Dict[str, int] = {}
        for s in suppliers:
            sup_cats = s.get("business_categories") or [s.get("business_category") or "UNKNOWN"]
            for cat in sup_cats:
                cat_totals[cat] = cat_totals.get(cat, 0) + 1

        # Build supplier lookup
        supplier_lookup = {s["id"]: s for s in suppliers}

        # Group docs by category → unique submitters per category
        cat_submitters: Dict[str, set] = {}
        cat_docs: Dict[str, list] = {}
        for doc in sustainability_docs:
            sup = supplier_lookup.get(doc["supplier_id"])
            if not sup:
                continue
            sup_cats = sup.get("business_categories") or [sup.get("business_category") or "UNKNOWN"]
            for cat in sup_cats:
                if cat not in cat_submitters:
                    cat_submitters[cat] = set()
                    cat_docs[cat] = []
                cat_submitters[cat].add(doc["supplier_id"])
                cat_docs[cat].append({"supplier": sup, **doc})

        total_suppliers = len(suppliers)
        total_submitters = len({d["supplier_id"] for d in sustainability_docs if d["supplier_id"] in supplier_lookup})
        overall_rate = (total_submitters / total_suppliers * 100) if total_suppliers else 0.0

        by_category = []
        for cat, total in cat_totals.items():
            submitters = len(cat_submitters.get(cat, set()))
            rate = (submitters / total * 100) if total else 0.0
            by_category.append({
                "category": cat,
                "display": cat.replace("_", " ").title(),
                "total": total,
                "submitters": submitters,
                "rate": rate,
                "docs": sorted(
                    cat_docs.get(cat, []),
                    key=lambda d: (d["supplier"].get("company_name") or "").lower(),
                ),
            })

        # Sort: highest coverage first, then alpha
        by_category.sort(key=lambda x: (-x["rate"], x["display"].lower()))

        return {
            "total_suppliers": total_suppliers,
            "total_submitters": total_submitters,
            "overall_rate": overall_rate,
            "by_category": by_category,
        }

    # ──────────────────────────────────────────────────────────────────────
    # PDF report
    # ──────────────────────────────────────────────────────────────────────

    async def generate_pdf_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[List[SupplierStatus]] = None,
        category: Optional[List[BusinessCategory]] = None,
        location: Optional[str] = None,
        min_years: Optional[int] = None,
        max_years: Optional[int] = None,
        sort_by: str = "company_name",
        sort_order: str = "asc",
        include_summary: bool = True,
        include_sustainability: bool = True,
        include_supplier_list: bool = True,
    ) -> BytesIO:
        """
        Generate a PDF report of suppliers.
        """
        # Get filtered data
        suppliers = await self.get_filtered_suppliers(
            start_date, end_date, status, category, location, min_years, max_years,
            sort_by=sort_by, sort_order=sort_order,
        )

        # Create PDF buffer
        buffer = BytesIO()
        
        # Create document with landscape orientation for better table fit
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles with RTG branding
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066cc'),  # RTG Blue
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#004c99'),  # RTG Dark Blue
            spaceAfter=12,
            fontName='Helvetica-Bold',
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#4b5563'),
        )

        # Cell styles for wrapping table content
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Helvetica',
            leading=11,
            textColor=colors.HexColor('#1f2937'),
        )
        cat_cell_style = ParagraphStyle(
            'CatCellStyle',
            parent=styles['Normal'],
            fontSize=7.5,
            fontName='Helvetica',
            leading=10,
            textColor=colors.HexColor('#374151'),
        )
        kpi_style = ParagraphStyle(
            'KPIStyle',
            parent=styles['Normal'],
            fontSize=18,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#0066cc'),
            alignment=TA_CENTER,
            leading=22,
        )

        # Status colour map used throughout
        STATUS_COLORS = {
            'APPROVED': colors.HexColor('#16a34a'),
            'PENDING': colors.HexColor('#9ca3af'),
            'UNDER_REVIEW': colors.HexColor('#2563eb'),
            'COMPLIANCE_REQUIRED': colors.HexColor('#d97706'),
            'SUSPENDED': colors.HexColor('#dc2626'),
            'REJECTED': colors.HexColor('#7f1d1d'),
        }

        # Add title
        logo_path = os.path.join(os.path.dirname(__file__), '..', 'core', 'rtg-logo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=2*inch, height=0.6*inch, kind='proportional')
            elements.append(logo)
            elements.append(Spacer(1, 0.2 * inch))
        
        elements.append(Paragraph(f"{self.company_name}", title_style))
        elements.append(Paragraph("Supplier Details Report", heading_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Add report metadata
        report_info = [
            f"<b>Generated:</b> {get_cat_now().strftime('%B %d, %Y at %I:%M %p CAT')}",
            f"<b>Total Suppliers:</b> {len(suppliers)}",
        ]
        
        if start_date:
            report_info.append(f"<b>From:</b> {start_date.strftime('%B %d, %Y')}")
        if end_date:
            report_info.append(f"<b>To:</b> {end_date.strftime('%B %d, %Y')}")
        if status:
            report_info.append(f"<b>Status Filter:</b> {', '.join([s.value for s in status])}")
        if category:
            report_info.append(f"<b>Category Filter:</b> {', '.join([c.value.replace('_', ' ').title() for c in category])}")
        if location:
            report_info.append(f"<b>Location Filter:</b> {location}")
        
        for info in report_info:
            elements.append(Paragraph(info, normal_style))
        
        elements.append(Spacer(1, 0.3 * inch))

        # ── Executive KPI Banner ───────────────────────────────────────────
        if suppliers:
            _unique_cats: set = set()
            _total_years = 0.0
            _years_count = 0
            for _s in suppliers:
                for _c in (_s.get('business_categories') or [_s.get('business_category') or 'Unknown']):
                    _unique_cats.add(_c)
                _yib = _s.get('years_in_business')
                if _yib is not None:
                    try:
                        _total_years += float(_yib)
                        _years_count += 1
                    except (ValueError, TypeError):
                        pass
            _avg_years = f"{_total_years / _years_count:.1f}" if _years_count else "N/A"

            def _kpi_cell(value: str, label: str) -> Paragraph:
                return Paragraph(
                    f"<font size='20' color='#0066cc'><b>{value}</b></font><br/>"
                    f"<font size='8' color='#6b7280'>{label}</font>",
                    kpi_style,
                )

            _kpi_data = [[
                _kpi_cell(str(len(suppliers)), "Total Approved Suppliers"),
                _kpi_cell(str(len(_unique_cats)), "Active Supply Categories"),
                _kpi_cell(f"{_avg_years} yrs", "Avg. Business Tenure"),
                _kpi_cell(get_cat_now().strftime('%b %Y'), "Report Period"),
            ]]
            _kpi_tbl = Table(_kpi_data, colWidths=[2.7 * inch] * 4)
            _kpi_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f7ff')),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#0066cc')),
                ('LINEAFTER', (0, 0), (2, -1), 0.75, colors.HexColor('#bfdbfe')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 14),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
            ]))
            elements.append(_kpi_tbl)
            elements.append(Spacer(1, 0.3 * inch))

        # ── Summary section ────────────────────────────────────────────────
        if include_summary and suppliers:
            elements.append(Paragraph("Report Summary", heading_style))
            elements.append(Spacer(1, 0.2 * inch))

            _status_counts: Dict[str, int] = {}
            _cat_counts: Dict[str, int] = {}
            for _s in suppliers:
                _sv = _s.get('status', 'Unknown')
                _status_counts[_sv] = _status_counts.get(_sv, 0) + 1
                for _cv in (_s.get('business_categories') or [_s.get('business_category') or 'Unknown']):
                    _cat_counts[_cv] = _cat_counts.get(_cv, 0) + 1

            elements.append(Paragraph("Status Distribution", normal_style))
            elements.append(Spacer(1, 0.1 * inch))
            _st_data = [['Status', 'Suppliers', 'Percentage']]
            _st_status_vals = []
            for _sv, _cnt in sorted(_status_counts.items()):
                _st_data.append([_sv.upper(), str(_cnt), f"{_cnt/len(suppliers)*100:.1f}%"])
                _st_status_vals.append(_sv.upper())
            _st_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
            ]
            for _ri, _sv in enumerate(_st_status_vals, 1):
                _sc = STATUS_COLORS.get(_sv, colors.HexColor('#6b7280'))
                _st_style.append(('TEXTCOLOR', (0, _ri), (0, _ri), _sc))
                _st_style.append(('FONTNAME', (0, _ri), (0, _ri), 'Helvetica-Bold'))
            _st_tbl = Table(_st_data, colWidths=[3*inch, 2*inch, 1.5*inch])
            _st_tbl.setStyle(TableStyle(_st_style))
            elements.append(_st_tbl)
            elements.append(Spacer(1, 0.3 * inch))

            elements.append(Paragraph("Category Distribution", normal_style))
            elements.append(Spacer(1, 0.1 * inch))
            _cd_data = [['Business Category', 'Suppliers', 'Share', 'Coverage']]
            _cd_pcts = []
            _total_supplier_count = len(suppliers)
            for _cv, _cnt in sorted(_cat_counts.items(), key=lambda x: x[1], reverse=True):
                _pct = (_cnt / _total_supplier_count * 100) if _total_supplier_count else 0.0
                _filled = round(_pct / 10)
                _bar = ('█' * _filled) + ('░' * (10 - _filled))
                _cd_data.append([_cv.replace('_', ' ').title(), str(_cnt), f"{_pct:.1f}%", _bar])
                _cd_pcts.append(_pct)
            _cd_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('FONTNAME', (3, 1), (3, -1), 'Courier'),
                ('FONTSIZE', (3, 1), (3, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
            ]
            for _ri, _pct in enumerate(_cd_pcts, 1):
                if _pct >= 33:
                    _pct_c = colors.HexColor('#16a34a')
                elif _pct >= 15:
                    _pct_c = colors.HexColor('#d97706')
                else:
                    _pct_c = colors.HexColor('#dc2626')
                _cd_style.append(('TEXTCOLOR', (2, _ri), (2, _ri), _pct_c))
                _cd_style.append(('FONTNAME', (2, _ri), (2, _ri), 'Helvetica-Bold'))
                _cd_style.append(('TEXTCOLOR', (3, _ri), (3, _ri), _pct_c))
            _cd_tbl = Table(_cd_data, colWidths=[3.2*inch, 1.3*inch, 1.0*inch, 2.0*inch])
            _cd_tbl.setStyle(TableStyle(_cd_style))
            elements.append(_cd_tbl)
            elements.append(PageBreak())

        # ── Sustainability section (per-category) ──────────────────────────
        if include_sustainability:
            _enhanced_metrics = await self._build_enhanced_sustainability_metrics(suppliers)

            elements.append(Paragraph("Sustainability & ESG Metrics", heading_style))
            elements.append(Spacer(1, 0.15 * inch))

            elements.append(Paragraph(
                f"<b>Scope:</b> {_enhanced_metrics['total_suppliers']} APPROVED suppliers.",
                normal_style,
            ))
            elements.append(Spacer(1, 0.08 * inch))

            elements.append(Paragraph(
                "Enhanced ESG metrics are derived from current sustainability analytics (key-person profile and supplier ESG summary data).",
                normal_style,
            ))
            elements.append(Spacer(1, 0.15 * inch))

            _esg_rows = [[
                'Metric',
                'Count',
                'Percentage',
            ], [
                'Women-Owned / Led Suppliers',
                str(_enhanced_metrics['women_owned_count']),
                f"{_enhanced_metrics['women_owned_pct']:.1f}%",
            ], [
                'Youth-Owned / Led Suppliers',
                str(_enhanced_metrics['youth_owned_count']),
                f"{_enhanced_metrics['youth_owned_pct']:.1f}%",
            ], [
                'Small-Scale Farmers',
                str(_enhanced_metrics['small_scale_farmer_count']),
                f"{_enhanced_metrics['small_scale_farmer_pct']:.1f}%",
            ], [
                'Female Directors',
                str(_enhanced_metrics['female_directors']),
                f"{_enhanced_metrics['female_director_pct']:.1f}% of directors",
            ], [
                'Youth Directors',
                str(_enhanced_metrics['youth_directors']),
                f"{_enhanced_metrics['youth_director_pct']:.1f}% of directors",
            ], [
                'Total SMEs (Small + Medium)',
                str(_enhanced_metrics['sme_total']),
                f"{_enhanced_metrics['sme_pct']:.1f}%",
            ]]

            _esg_tbl = Table(_esg_rows, colWidths=[3.8*inch, 1.6*inch, 2.4*inch], repeatRows=1)
            _esg_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfeff')]),
            ]))
            elements.append(_esg_tbl)
            elements.append(Spacer(1, 0.2 * inch))

            elements.append(Paragraph("Business Size / SME Breakdown", normal_style))
            elements.append(Spacer(1, 0.08 * inch))
            _size_rows = [['Business Size', 'Suppliers', 'Percentage']]
            for _size in _enhanced_metrics["business_size_distribution"]:
                _size_rows.append([
                    _size["business_size"].title(),
                    str(_size["supplier_count"]),
                    f"{_size['pct']:.1f}%",
                ])
            _size_tbl = Table(_size_rows, colWidths=[3.2*inch, 1.6*inch, 1.8*inch], repeatRows=1)
            _size_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
            ]))
            elements.append(_size_tbl)
            elements.append(Spacer(1, 0.25 * inch))

            # ── Category Compliance ───────────────────────────────────────
            _cat_compliance = await db.get_category_compliance_stats(status="APPROVED")
            _cat_compliance = [r for r in _cat_compliance if r.get("total_suppliers", 0) > 0]

            elements.append(Paragraph("Category Compliance Overview", normal_style))
            elements.append(Spacer(1, 0.08 * inch))
            _cc_data = [["Business Category", "Total", "Mand. Met", "Mand. Missing",
                         "Full Compliance", "Med Risk", "High Risk", "Pending", "Compliance %"]]
            for _r in _cat_compliance:
                _cc_data.append([
                    _r["category"].replace("_", " ").title()[:30],
                    str(_r["total_suppliers"]),
                    str(_r["mandatory_met_count"]),
                    str(_r["mandatory_missing_count"]),
                    str(_r["full_compliance_count"]),
                    str(_r["medium_risk_count"]),
                    str(_r["high_risk_count"]),
                    str(_r["pending_count"]),
                    f"{_r['full_compliance_pct']:.1f}%",
                ])
            if len(_cc_data) > 1:
                _cc_tbl = Table(_cc_data, colWidths=[2.1*inch, 0.55*inch, 0.7*inch, 0.9*inch,
                                                      0.85*inch, 0.65*inch, 0.65*inch, 0.55*inch, 0.8*inch], repeatRows=1)
                _cc_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1d4ed8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')]),
                ]
                for _ri, _r in enumerate(_cat_compliance, start=1):
                    if _r.get("high_risk_count", 0) > 0:
                        _cc_style.append(('TEXTCOLOR', (6, _ri), (6, _ri), colors.HexColor('#dc2626')))
                    if _r.get("mandatory_missing_count", 0) > 0:
                        _cc_style.append(('TEXTCOLOR', (3, _ri), (3, _ri), colors.HexColor('#dc2626')))
                _cc_tbl.setStyle(TableStyle(_cc_style))
                elements.append(_cc_tbl)
            else:
                elements.append(Paragraph("No category compliance data available.", normal_style))
            elements.append(Spacer(1, 0.25 * inch))

            # ── Document Type Statistics ──────────────────────────────────
            _doc_stats = await db.get_document_type_stats()
            if _doc_stats:
                elements.append(Paragraph("Document Coverage & Verification Status", normal_style))
                elements.append(Spacer(1, 0.08 * inch))
                _ds_data = [["Document Type", "Suppliers", "Total Uploaded", "Verified", "Pending", "Rejected"]]
                for _r in sorted(_doc_stats, key=lambda r: r.get("supplier_count", 0), reverse=True):
                    _ds_data.append([
                        (_r.get("document_type") or "").replace("_", " ").title()[:35],
                        str(_r.get("supplier_count", 0)),
                        str(_r.get("total_uploads", 0)),
                        str(_r.get("verified_count", 0)),
                        str(_r.get("pending_count", 0)),
                        str(_r.get("rejected_count", 0)),
                    ])
                _ds_tbl = Table(_ds_data, colWidths=[3.0*inch, 0.9*inch, 1.0*inch, 0.9*inch, 0.9*inch, 0.9*inch], repeatRows=1)
                _ds_tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfeff')]),
                ]))
                elements.append(_ds_tbl)
                elements.append(Spacer(1, 0.2 * inch))

            elements.append(PageBreak())

        # ── Supplier list ─────────────────────────────────────────────────
        if include_supplier_list:
            elements.append(Paragraph("Approved Supplier Directory", heading_style))
            elements.append(Spacer(1, 0.15 * inch))

            _hdr_para = ParagraphStyle(
                'HdrCell', fontSize=8, fontName='Helvetica-Bold',
                textColor=colors.white, alignment=TA_CENTER, leading=10,
            )
            table_data = [[
                Paragraph('Company Name', _hdr_para),
                Paragraph('Business Categories', _hdr_para),
                Paragraph('Location', _hdr_para),
                Paragraph('Contact Person', _hdr_para),
                Paragraph('Email', _hdr_para),
                Paragraph('Status', _hdr_para),
                Paragraph('Yrs', _hdr_para),
                Paragraph('Registered', _hdr_para),
            ]]
            _tbl_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7ff')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
            ]
            for i, supplier in enumerate(suppliers, 1):
                cats = supplier.get('business_categories') or [supplier.get('business_category') or 'N/A']
                cat_text = '\n'.join(c.replace('_', ' ').title() for c in cats)
                status = (supplier.get('status') or 'N/A').upper()
                row = [
                    Paragraph(supplier.get('company_name', 'N/A'), cell_style),
                    Paragraph(cat_text, cat_cell_style),
                    f"{supplier.get('city', '') or ''}, {supplier.get('country', '') or ''}"[:22],
                    (supplier.get('contact_person_name', '') or '')[:20],
                    (supplier.get('email', '') or '')[:28],
                    status,
                    str(supplier.get('years_in_business', 'N/A')),
                    format_cat_datetime(supplier.get('created_at'), '%Y-%m-%d') if supplier.get('created_at') else 'N/A',
                ]
                table_data.append(row)
                status_color = STATUS_COLORS.get(status, colors.HexColor('#6b7280'))
                _tbl_style.append(('TEXTCOLOR', (5, i), (5, i), status_color))
                _tbl_style.append(('FONTNAME', (5, i), (5, i), 'Helvetica-Bold'))

            _col_widths = [1.85*inch, 2.3*inch, 1.0*inch, 1.1*inch, 1.85*inch, 0.75*inch, 0.5*inch, 0.8*inch]
            table = Table(table_data, repeatRows=1, colWidths=_col_widths)
            table.setStyle(TableStyle(_tbl_style))
            elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    # ──────────────────────────────────────────────────────────────────────
    # Excel report
    # ──────────────────────────────────────────────────────────────────────

    async def generate_excel_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[List[SupplierStatus]] = None,
        category: Optional[List[BusinessCategory]] = None,
        location: Optional[str] = None,
        min_years: Optional[int] = None,
        max_years: Optional[int] = None,
        sort_by: str = "company_name",
        sort_order: str = "asc",
        include_summary: bool = True,
        include_sustainability: bool = True,
        include_supplier_list: bool = True,
    ) -> BytesIO:
        """
        Generate an Excel report of suppliers.
        Sheets created depend on the include_* flags:
          - Summary            (if include_summary)
          - Sustainability Report (if include_sustainability)
          - Supplier Details   (if include_supplier_list)
        """
        # Get filtered data
        suppliers = await self.get_filtered_suppliers(
            start_date, end_date, status, category, location, min_years, max_years,
            sort_by=sort_by, sort_order=sort_order,
        )

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # --- Shared styling ---
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Pre-compute ESG metrics
        _xl_enhanced_data = await self._build_enhanced_sustainability_metrics(suppliers)

        # ===== Sheet 1: Summary =====
        if include_summary:
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Report Summary").font = Font(bold=True, size=16)
            ws_summary.cell(row=2, column=1, value=f"Generated: {get_cat_now().strftime('%B %d, %Y at %I:%M %p CAT')}")
            ws_summary.cell(row=3, column=1, value=f"Total Suppliers: {len(suppliers)}")

            row_offset = 5

            if any([start_date, end_date, status, category, location]):
                ws_summary.cell(row=row_offset, column=1, value="Filters Applied:").font = Font(bold=True)
                row_offset += 1
                if start_date:
                    ws_summary.cell(row=row_offset, column=1, value=f"From: {start_date.strftime('%B %d, %Y')}")
                    row_offset += 1
                if end_date:
                    ws_summary.cell(row=row_offset, column=1, value=f"To: {end_date.strftime('%B %d, %Y')}")
                    row_offset += 1
                if status:
                    ws_summary.cell(row=row_offset, column=1, value=f"Status: {', '.join([s.value for s in status])}")
                    row_offset += 1
                if category:
                    ws_summary.cell(row=row_offset, column=1, value=f"Category: {', '.join([c.value for c in category])}")
                    row_offset += 1
                if location:
                    ws_summary.cell(row=row_offset, column=1, value=f"Location: {location}")
                    row_offset += 1
                row_offset += 1

            ws_summary.cell(row=row_offset, column=1, value="Status Distribution").font = Font(bold=True, size=12)
            row_offset += 1
            status_counts: dict = {}
            for supplier in suppliers:
                sv = supplier.get('status', 'Unknown')
                status_counts[sv] = status_counts.get(sv, 0) + 1
            for col, hdr in enumerate(['Status', 'Count', 'Percentage'], 1):
                cell = ws_summary.cell(row=row_offset, column=col, value=hdr)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row_offset += 1
            for sv, count in sorted(status_counts.items()):
                pct = (count / len(suppliers)) * 100 if suppliers else 0
                ws_summary.cell(row=row_offset, column=1, value=sv.upper()).border = border
                ws_summary.cell(row=row_offset, column=2, value=count).border = border
                ws_summary.cell(row=row_offset, column=3, value=f"{pct:.1f}%").border = border
                row_offset += 1
            row_offset += 2

            ws_summary.cell(row=row_offset, column=1, value="Category Distribution").font = Font(bold=True, size=12)
            row_offset += 1
            category_counts: dict = {}
            for supplier in suppliers:
                for cv in (supplier.get('business_categories') or [supplier.get('business_category') or 'Unknown']):
                    category_counts[cv] = category_counts.get(cv, 0) + 1
            for col, hdr in enumerate(['Business Category', 'Count', 'Percentage'], 1):
                cell = ws_summary.cell(row=row_offset, column=col, value=hdr)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row_offset += 1
            for cv, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
                pct = (count / len(suppliers)) * 100 if suppliers else 0
                ws_summary.cell(row=row_offset, column=1, value=cv.replace('_', ' ').title()).border = border
                ws_summary.cell(row=row_offset, column=2, value=count).border = border
                ws_summary.cell(row=row_offset, column=3, value=f"{pct:.1f}%").border = border
                row_offset += 1
            row_offset += 2

            ws_summary.cell(row=row_offset, column=1, value="Sustainability & ESG Metrics").font = Font(bold=True, size=12)
            row_offset += 1
            ws_summary.cell(
                row=row_offset, column=1,
                value=(
                    f"Women-Owned/Led: {_xl_enhanced_data['women_owned_count']} "
                    f"({_xl_enhanced_data['women_owned_pct']:.1f}%) | "
                    f"Youth-Owned/Led: {_xl_enhanced_data['youth_owned_count']} "
                    f"({_xl_enhanced_data['youth_owned_pct']:.1f}%)"
                ),
            )
            row_offset += 1
            ws_summary.cell(row=row_offset, column=1,
                            value=(
                                f"SMEs (Small + Medium): {_xl_enhanced_data['sme_total']} "
                                f"({_xl_enhanced_data['sme_pct']:.1f}%) | "
                                f"Small-Scale Farmers: {_xl_enhanced_data['small_scale_farmer_count']} "
                                f"({_xl_enhanced_data['small_scale_farmer_pct']:.1f}%)"
                            ))
            row_offset += 1
            ws_summary.cell(
                row=row_offset,
                column=1,
                value="Full ESG + certification breakdown available in the 'Sustainability Report' sheet.",
            )
            for col in range(1, 4):
                ws_summary.column_dimensions[get_column_letter(col)].width = 40

        # ===== Sheet 2: Sustainability & ESG Metrics =====
        if include_sustainability:
            sustain_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")

            ws_sustain = wb.create_sheet("Sustainability Report")
            ws_sustain.cell(row=1, column=1, value="Sustainability & ESG Metrics").font = Font(bold=True, size=16, color="0F766E")
            ws_sustain.cell(row=2, column=1, value=f"Generated: {get_cat_now().strftime('%B %d, %Y at %I:%M %p CAT')}")
            ws_sustain.cell(row=3, column=1,
                            value=(
                                f"Women-Owned/Led: {_xl_enhanced_data['women_owned_count']} ({_xl_enhanced_data['women_owned_pct']:.1f}%), "
                                f"Youth-Owned/Led: {_xl_enhanced_data['youth_owned_count']} ({_xl_enhanced_data['youth_owned_pct']:.1f}%), "
                                f"SMEs: {_xl_enhanced_data['sme_total']} ({_xl_enhanced_data['sme_pct']:.1f}%)"
                            ))
            c4 = ws_sustain.cell(row=4, column=1,
                                 value="Enhanced ESG metrics align with the Sustainability dashboard logic.")
            c4.font = Font(italic=True, color="6B7280")
            ws_sustain.cell(
                row=5,
                column=1,
                value=f"Scope: {_xl_enhanced_data['total_suppliers']} APPROVED suppliers.",
            ).font = Font(bold=True, color="1F2937")

            # Enhanced ESG metric table
            ws_sustain.cell(row=7, column=1, value="Enhanced ESG Snapshot").font = Font(bold=True, size=12)
            for _ci, _hdr in enumerate(['Metric', 'Count', 'Percentage'], 1):
                cell = ws_sustain.cell(row=8, column=_ci, value=_hdr)
                cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            _esg_table_rows = [
                ["Women-Owned / Led Suppliers", _xl_enhanced_data["women_owned_count"], f"{_xl_enhanced_data['women_owned_pct']:.1f}%"],
                ["Youth-Owned / Led Suppliers", _xl_enhanced_data["youth_owned_count"], f"{_xl_enhanced_data['youth_owned_pct']:.1f}%"],
                ["Small-Scale Farmers", _xl_enhanced_data["small_scale_farmer_count"], f"{_xl_enhanced_data['small_scale_farmer_pct']:.1f}%"],
                ["Female Directors", _xl_enhanced_data["female_directors"], f"{_xl_enhanced_data['female_director_pct']:.1f}% of directors"],
                ["Youth Directors", _xl_enhanced_data["youth_directors"], f"{_xl_enhanced_data['youth_director_pct']:.1f}% of directors"],
                ["Total SMEs (Small + Medium)", _xl_enhanced_data["sme_total"], f"{_xl_enhanced_data['sme_pct']:.1f}%"],
            ]
            for _ri, _row in enumerate(_esg_table_rows, start=9):
                for _ci, _val in enumerate(_row, 1):
                    cell = ws_sustain.cell(row=_ri, column=_ci, value=_val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if _ri % 2 == 0:
                        cell.fill = PatternFill(start_color="ECFEFF", end_color="ECFEFF", fill_type="solid")

            # Business size distribution table
            _size_start = 8 + len(_esg_table_rows) + 2
            ws_sustain.cell(row=_size_start, column=1, value="Business Size / SME Breakdown").font = Font(bold=True, size=12)
            for _ci, _hdr in enumerate(['Business Size', 'Suppliers', 'Percentage'], 1):
                cell = ws_sustain.cell(row=_size_start + 1, column=_ci, value=_hdr)
                cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            for _ri, _row in enumerate(_xl_enhanced_data["business_size_distribution"], start=_size_start + 2):
                ws_sustain.cell(row=_ri, column=1, value=str(_row["business_size"]).title()).border = border
                ws_sustain.cell(row=_ri, column=2, value=_row["supplier_count"]).border = border
                ws_sustain.cell(row=_ri, column=3, value=f"{_row['pct']:.1f}%").border = border
                if _ri % 2 == 0:
                    for _ci in (1, 2, 3):
                        ws_sustain.cell(row=_ri, column=_ci).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

            # ── Category Compliance ───────────────────────────────────────────
            _xl_cat_compl = await db.get_category_compliance_stats(status="APPROVED")
            _xl_cat_compl = [r for r in _xl_cat_compl if r.get("total_suppliers", 0) > 0]

            _compl_start = _size_start + 2 + len(_xl_enhanced_data["business_size_distribution"]) + 2
            ws_sustain.cell(row=_compl_start, column=1, value="Category Compliance Overview").font = Font(bold=True, size=12)
            _cc_hdrs = ["Business Category", "Total", "Mand. Met", "Mand. Missing",
                        "Full Compliance", "Med Risk", "High Risk", "Pending", "Compliance %"]
            for _ci, _hdr in enumerate(_cc_hdrs, 1):
                cell = ws_sustain.cell(row=_compl_start + 1, column=_ci, value=_hdr)
                cell.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            for _ri, _r in enumerate(_xl_cat_compl, start=_compl_start + 2):
                _vals = [
                    _r["category"].replace("_", " ").title(),
                    _r["total_suppliers"], _r["mandatory_met_count"], _r["mandatory_missing_count"],
                    _r["full_compliance_count"], _r["medium_risk_count"], _r["high_risk_count"],
                    _r["pending_count"], f"{_r['full_compliance_pct']:.1f}%",
                ]
                for _ci, _val in enumerate(_vals, 1):
                    cell = ws_sustain.cell(row=_ri, column=_ci, value=_val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                    if _ri % 2 == 0:
                        cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
                    if _ci == 4 and isinstance(_val, int) and _val > 0:
                        cell.font = Font(color="DC2626", bold=True)
                    if _ci == 7 and isinstance(_val, int) and _val > 0:
                        cell.font = Font(color="DC2626")

            # ── Document Type Stats ───────────────────────────────────────────
            _xl_doc_stats = await db.get_document_type_stats()
            _doc_start = _compl_start + 2 + len(_xl_cat_compl) + 2
            ws_sustain.cell(row=_doc_start, column=1, value="Document Coverage & Verification Status").font = Font(bold=True, size=12)
            _ds_hdrs = ["Document Type", "Suppliers", "Total Uploaded", "Verified", "Pending", "Rejected"]
            for _ci, _hdr in enumerate(_ds_hdrs, 1):
                cell = ws_sustain.cell(row=_doc_start + 1, column=_ci, value=_hdr)
                cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            for _ri, _r in enumerate(sorted(_xl_doc_stats, key=lambda r: r.get("supplier_count", 0), reverse=True), start=_doc_start + 2):
                _vals = [
                    (_r.get("document_type") or "").replace("_", " ").title(),
                    _r.get("supplier_count", 0), _r.get("total_uploads", 0),
                    _r.get("verified_count", 0), _r.get("pending_count", 0), _r.get("rejected_count", 0),
                ]
                for _ci, _val in enumerate(_vals, 1):
                    cell = ws_sustain.cell(row=_ri, column=_ci, value=_val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if _ri % 2 == 0:
                        cell.fill = PatternFill(start_color="ECFEFF", end_color="ECFEFF", fill_type="solid")

            for _ci, _w in enumerate([35, 12, 14, 14, 14, 14, 14, 12, 14], 1):
                ws_sustain.column_dimensions[get_column_letter(_ci)].width = _w

        # ===== Sheet 3: Supplier Details =====
        if include_supplier_list:
            ws_details = wb.create_sheet("Supplier Details")
            detail_headers = [
                'Company Name', 'Business Category', 'Registration Number', 'Tax ID',
                'Years in Business', 'Website', 'Contact Person', 'Title',
                'Email', 'Phone', 'Street Address', 'City', 'State/Province',
                'Postal Code', 'Country', 'Status', 'Created Date', 'Submitted Date', 'Updated Date',
            ]
            for col_num, hdr in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col_num, value=hdr)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            for row_num, supplier in enumerate(suppliers, 2):
                data = [
                    supplier.get('company_name', ''),
                    ', '.join(c.replace('_', ' ').title() for c in (supplier.get('business_categories') or [supplier.get('business_category') or ''])),
                    supplier.get('registration_number', ''),
                    supplier.get('tax_id', ''),
                    supplier.get('years_in_business', ''),
                    supplier.get('website', ''),
                    supplier.get('contact_person_name', ''),
                    supplier.get('contact_person_title', ''),
                    supplier.get('email', ''),
                    supplier.get('phone', ''),
                    supplier.get('street_address', ''),
                    supplier.get('city', ''),
                    supplier.get('state_province', ''),
                    supplier.get('postal_code', ''),
                    supplier.get('country', ''),
                    (supplier.get('status') or '').upper(),
                    self._format_date(supplier.get('created_at')),
                    self._format_date(supplier.get('submitted_at')),
                    self._format_date(supplier.get('updated_at')),
                ]
                for col_num, value in enumerate(data, 1):
                    cell = ws_details.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            for col_num, hdr in enumerate(detail_headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(hdr)
                for row in ws_details.iter_rows(min_col=col_num, max_col=col_num, min_row=2):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws_details.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _format_date(self, date_str: Optional[str]) -> str:
        """Format ISO date string to readable format."""
        if not date_str:
            return ''
        try:
            return format_cat_datetime(date_str, '%Y-%m-%d %H:%M')
        except:
            return date_str


# Global instance
report_service = ReportService()
