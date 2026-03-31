"""
Sustainability & ESG reporting API routes.

Exposes aggregated metrics from the four compliance views:
  - vw_esg_supplier_summary
  - vw_category_compliance
  - vw_business_size_distribution
  - vw_document_type_stats

Also provides filtered supplier list and CSV/Excel downloads.
"""

from io import BytesIO, StringIO
import asyncio
import csv
import time
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi import status as http_status
from fastapi.responses import StreamingResponse

from ...db.supabase import db
from ...api.deps import get_current_admin
from ...core.timezone import get_cat_now, get_cat_timestamp_str
from ...core.logger import logger
from ...models.enums import SupplierStatus

router = APIRouter(prefix="/sustainability", tags=["Sustainability"])

APPROVED_STATUS_SCOPE = ["APPROVED", "COMPLIANCE_REQUIRED", "SUSPENDED"]
MAINTENANCE_INTERVAL_SECONDS = 300
OVERVIEW_CACHE_TTL_SECONDS = 300  # 5 minutes — reads from DB views are fast

_last_maintenance_run_ts = 0.0
_maintenance_lock = asyncio.Lock()
_overview_cache: dict[str, tuple[float, dict]] = {}
_overview_cache_lock = asyncio.Lock()


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_pct(part: int, total: int) -> float:
    return round(100 * part / total, 1) if total else 0.0


def _derive_esg_booleans_from_counts(row: dict) -> tuple[bool, bool]:
    """Derive ownership/leadership flags using >50% threshold rules."""
    total_key_persons = row.get("key_person_count") or 0
    female_count = row.get("female_director_count") or 0
    youth_count = row.get("youth_director_count") or 0

    if total_key_persons <= 0:
        return False, False

    return (female_count / total_key_persons) > 0.5, (youth_count / total_key_persons) > 0.5


async def _run_sustainability_maintenance_if_due(
    statuses: list[str],
    force_refresh: bool = False,
) -> None:
    """Throttle heavy maintenance work to avoid slowing every dashboard request."""
    global _last_maintenance_run_ts

    now = time.time()
    if not force_refresh and now - _last_maintenance_run_ts < MAINTENANCE_INTERVAL_SECONDS:
        return

    async with _maintenance_lock:
        # Re-check inside lock so concurrent requests do not duplicate work.
        now = time.time()
        if not force_refresh and now - _last_maintenance_run_ts < MAINTENANCE_INTERVAL_SECONDS:
            return

        try:
            await db.backfill_supplier_categories_from_primary(statuses=statuses)
        except Exception as backfill_err:
            logger.warning("supplier category backfill skipped: %s", backfill_err)

        try:
            await db.recompute_category_compliance_for_suppliers(statuses=statuses)
        except Exception as compliance_err:
            logger.warning("category compliance refresh skipped: %s", compliance_err)

        _last_maintenance_run_ts = time.time()


def _build_overview_cache_key(
    country: Optional[str],
    business_size: Optional[str],
    status: Optional[str],
) -> str:
    return f"country={country or ''}|business_size={business_size or ''}|status={status or ''}"


async def _get_cached_overview(cache_key: str) -> Optional[dict]:
    async with _overview_cache_lock:
        cached = _overview_cache.get(cache_key)
        if not cached:
            return None

        cached_at, payload = cached
        if time.time() - cached_at > OVERVIEW_CACHE_TTL_SECONDS:
            _overview_cache.pop(cache_key, None)
            return None

        return payload


async def _set_cached_overview(cache_key: str, payload: dict) -> None:
    async with _overview_cache_lock:
        _overview_cache[cache_key] = (time.time(), payload)


# ── 1. Overview (all KPIs in one call) ──────────────────────────────────────

@router.get("/overview", summary="Full sustainability KPI snapshot")
async def get_sustainability_overview(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None, description="Supplier status filter"),
    refresh_stats: bool = Query(
        False,
        description="Force a one-time maintenance refresh before reading stats",
    ),
    current_admin: dict = Depends(get_current_admin),
):
    """
    Returns all sustainability metrics needed to render the dashboard:
      - ESG headline numbers
      - Business-size breakdown
      - Category compliance distribution
      - Document type upload stats
    """
    approved_scope = ["APPROVED", "COMPLIANCE_REQUIRED", "SUSPENDED"]
    maintenance_statuses = [
        SupplierStatus.INCOMPLETE.value,
        SupplierStatus.SUBMITTED.value,
        SupplierStatus.UNDER_REVIEW.value,
        SupplierStatus.NEED_MORE_INFO.value,
        SupplierStatus.APPROVED.value,
        SupplierStatus.COMPLIANCE_REQUIRED.value,
        SupplierStatus.SUSPENDED.value,
    ]
    status_filter = status if status else None
    cache_key = _build_overview_cache_key(country, business_size, status_filter)

    if refresh_stats:
        # Explicit refresh: invalidate the cache entry and await maintenance so the
        # response reflects freshly recomputed compliance data.
        async with _overview_cache_lock:
            _overview_cache.pop(cache_key, None)
        await _run_sustainability_maintenance_if_due(
            statuses=maintenance_statuses,
            force_refresh=True,
        )
    else:
        # --- Cache check FIRST ---
        cached_payload = await _get_cached_overview(cache_key)
        if cached_payload is not None:
            return cached_payload
        # Cache miss: proceed to DB reads immediately.
        # Do NOT fire background maintenance — that floods Supabase with 900+ requests
        # and makes the backend unresponsive. Use refresh_stats=true to recompute.

    esg_task = db.get_esg_summary(
        country=country,
        business_size=business_size,
        status=status_filter,
        columns="status,is_small_scale_farmer,key_person_count,female_director_count,youth_director_count",
    )
    size_task = db.get_business_size_distribution()
    cat_task = db.get_category_compliance_stats(
        country=country,
        business_size=business_size,
        status=status_filter,
    )
    doc_task = db.get_document_type_stats()

    esg_rows, size_rows, cat_rows, doc_rows = await asyncio.gather(
        esg_task,
        size_task,
        cat_task,
        doc_task,
    )
    if status_filter is None:
        esg_rows = [r for r in esg_rows if (r.get("status") or "") in APPROVED_STATUS_SCOPE]

    total = len(esg_rows)
    status_breakdown = {
        "approved": sum(1 for r in esg_rows if (r.get("status") or "") == "APPROVED"),
        "compliance_required": sum(1 for r in esg_rows if (r.get("status") or "") == "COMPLIANCE_REQUIRED"),
        "suspended": sum(1 for r in esg_rows if (r.get("status") or "") == "SUSPENDED"),
    }
    # Use key-person counts as source-of-truth for ownership/leadership metrics.
    derived_flags = [_derive_esg_booleans_from_counts(r) for r in esg_rows]
    women_owned = sum(1 for women_owned_flag, _ in derived_flags if women_owned_flag)
    youth_owned = sum(1 for _, youth_owned_flag in derived_flags if youth_owned_flag)
    farmers = sum(1 for r in esg_rows if r.get("is_small_scale_farmer"))
    female_directors = sum(r.get("female_director_count") or 0 for r in esg_rows)
    youth_directors = sum(r.get("youth_director_count") or 0 for r in esg_rows)
    total_directors = sum(r.get("key_person_count") or 0 for r in esg_rows)

    payload = {
        "total_suppliers": total,
        "supplier_status_breakdown": status_breakdown,
        "esg": {
            "women_owned_count": women_owned,
            "women_owned_pct": _safe_pct(women_owned, total),
            "youth_owned_count": youth_owned,
            "youth_owned_pct": _safe_pct(youth_owned, total),
            "small_scale_farmer_count": farmers,
            "small_scale_farmer_pct": _safe_pct(farmers, total),
            "female_directors": female_directors,
            "youth_directors": youth_directors,
            "total_directors": total_directors,
            "female_director_pct": _safe_pct(female_directors, total_directors),
            "youth_director_pct": _safe_pct(youth_directors, total_directors),
        },
        "business_size_distribution": size_rows,
        "category_compliance": cat_rows,
        "document_type_stats": doc_rows,
    }

    await _set_cached_overview(cache_key, payload)
    return payload


# ── 2. ESG supplier detail list ───────────────────────────────────────────────

@router.get("/suppliers", summary="Filtered supplier list for ESG reporting")
async def get_sustainability_suppliers(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    esg_women_owned: Optional[bool] = Query(None),
    esg_youth_owned: Optional[bool] = Query(None),
    is_small_scale_farmer: Optional[bool] = Query(None),
    current_admin: dict = Depends(get_current_admin),
):
    rows = await db.get_sustainability_supplier_list(
        country=country,
        business_size=business_size,
        status=status,
        esg_women_owned=esg_women_owned,
        esg_youth_owned=esg_youth_owned,
        is_small_scale_farmer=is_small_scale_farmer,
    )
    return {"suppliers": rows, "total": len(rows)}


@router.get("/compliance-audit", summary="Supplier-by-supplier category compliance audit")
async def get_compliance_audit(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    supplier_id: Optional[str] = Query(None),
    limit: int = Query(500, ge=1, le=2000),
    refresh_stats: bool = Query(
        False,
        description="Force a maintenance refresh before generating audit rows",
    ),
    current_admin: dict = Depends(get_current_admin),
):
    """Return per-supplier category compliance details for diagnostics and reconciliation."""
    maintenance_statuses = [
        SupplierStatus.INCOMPLETE.value,
        SupplierStatus.SUBMITTED.value,
        SupplierStatus.UNDER_REVIEW.value,
        SupplierStatus.NEED_MORE_INFO.value,
        SupplierStatus.APPROVED.value,
        SupplierStatus.COMPLIANCE_REQUIRED.value,
        SupplierStatus.SUSPENDED.value,
    ]

    await _run_sustainability_maintenance_if_due(
        statuses=maintenance_statuses,
        force_refresh=refresh_stats,
    )

    supplier_query = db.client.table("suppliers").select(
        "id,company_name,status,business_category,country,business_size"
    )
    if supplier_id:
        supplier_query = supplier_query.eq("id", supplier_id)
    if country:
        supplier_query = supplier_query.ilike("country", f"%{country}%")
    if business_size:
        supplier_query = supplier_query.eq("business_size", business_size)
    if status:
        supplier_query = supplier_query.eq("status", status)

    suppliers_result = supplier_query.limit(limit).execute()
    suppliers = suppliers_result.data or []
    supplier_ids = [row.get("id") for row in suppliers if row.get("id")]

    categories_by_supplier: dict[str, list[dict]] = {}
    if supplier_ids:
        category_rows_result = (
            db.client.table("supplier_categories")
            .select("supplier_id,category,compliance_status,compliance_checked_at")
            .in_("supplier_id", supplier_ids)
            .execute()
        )
        for row in (category_rows_result.data or []):
            sid = row.get("supplier_id")
            if not sid:
                continue
            categories_by_supplier.setdefault(sid, []).append(row)

    audit_rows = []
    summary = {
        "suppliers": len(suppliers),
        "fully_compliant": 0,
        "partial": 0,
        "non_compliant": 0,
        "missing_categories": 0,
    }

    for supplier in suppliers:
        sid = supplier.get("id")
        category_rows = categories_by_supplier.get(sid, [])

        full_count = sum(1 for row in category_rows if (row.get("compliance_status") or "").upper() == "FULL_COMPLIANCE")
        medium_count = sum(1 for row in category_rows if (row.get("compliance_status") or "").upper() == "MEDIUM_RISK")
        high_count = sum(1 for row in category_rows if (row.get("compliance_status") or "").upper() == "HIGH_RISK")
        pending_count = sum(1 for row in category_rows if (row.get("compliance_status") or "").upper() == "PENDING")
        excluded_count = sum(1 for row in category_rows if (row.get("compliance_status") or "").upper() == "EXCLUDED")

        mandatory_met = full_count + medium_count
        mandatory_missing = high_count + pending_count
        non_excluded_total = len(category_rows) - excluded_count
        allowed_categories = sorted(
            [
                (row.get("category") or "")
                for row in category_rows
                if (row.get("compliance_status") or "").upper() in ("FULL_COMPLIANCE", "MEDIUM_RISK")
                and (row.get("category") or "")
            ]
        )
        blocked_categories = sorted(
            [
                (row.get("category") or "")
                for row in category_rows
                if (row.get("compliance_status") or "").upper() in ("HIGH_RISK", "PENDING")
                and (row.get("category") or "")
            ]
        )

        if len(category_rows) == 0:
            portfolio_status = "NO_CATEGORY_ASSIGNMENT"
            summary["missing_categories"] += 1
        elif mandatory_missing == 0 and full_count == non_excluded_total:
            portfolio_status = "FULLY_COMPLIANT"
            summary["fully_compliant"] += 1
        elif mandatory_missing == 0:
            portfolio_status = "MANDATORY_MET_PENDING_PREFERRED"
            summary["fully_compliant"] += 1
        elif mandatory_met > 0:
            portfolio_status = "PARTIALLY_COMPLIANT"
            summary["partial"] += 1
        else:
            portfolio_status = "NON_COMPLIANT"
            summary["non_compliant"] += 1

        audit_rows.append(
            {
                "supplier_id": sid,
                "company_name": supplier.get("company_name"),
                "supplier_status": supplier.get("status"),
                "primary_category": supplier.get("business_category"),
                "country": supplier.get("country"),
                "business_size": supplier.get("business_size"),
                "category_count": len(category_rows),
                "mandatory_met_categories": mandatory_met,
                "mandatory_missing_categories": mandatory_missing,
                "allowed_categories": allowed_categories,
                "blocked_categories": blocked_categories,
                "full_compliance_categories": full_count,
                "medium_risk_categories": medium_count,
                "high_risk_categories": high_count,
                "pending_categories": pending_count,
                "excluded_categories": excluded_count,
                "portfolio_status": portfolio_status,
                "categories": sorted(
                    [
                        {
                            "category": row.get("category"),
                            "compliance_status": row.get("compliance_status"),
                            "compliance_checked_at": row.get("compliance_checked_at"),
                        }
                        for row in category_rows
                    ],
                    key=lambda item: item.get("category") or "",
                ),
            }
        )

    return {
        "summary": summary,
        "rows": audit_rows,
        "total": len(audit_rows),
    }


# ── 3. CSV download ──────────────────────────────────────────────────────────

@router.get("/export/csv", summary="Download ESG supplier list as CSV")
async def export_sustainability_csv(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    esg_women_owned: Optional[bool] = Query(None),
    esg_youth_owned: Optional[bool] = Query(None),
    is_small_scale_farmer: Optional[bool] = Query(None),
    current_admin: dict = Depends(get_current_admin),
):
    rows = await db.get_sustainability_supplier_list(
        country=country,
        business_size=business_size,
        status=status,
        esg_women_owned=esg_women_owned,
        esg_youth_owned=esg_youth_owned,
        is_small_scale_farmer=is_small_scale_farmer,
    )

    output = StringIO()
    fieldnames = [
        "company_name", "country", "status", "supplier_type",
        "business_size", "employee_count", "is_small_scale_farmer",
        "esg_women_owned", "esg_youth_owned",
        "business_categories", "years_in_business", "created_at",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        # Flatten list fields
        if isinstance(row.get("business_categories"), list):
            row["business_categories"] = ", ".join(row["business_categories"])
        writer.writerow(row)

    filename = f"rtg_sustainability_report_{get_cat_timestamp_str()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 4. Excel download (comprehensive multi-sheet report) ─────────────────────

@router.get("/export/excel", summary="Download comprehensive sustainability Excel report")
async def export_sustainability_excel(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_admin: dict = Depends(get_current_admin),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is required for Excel export. Run: pip install openpyxl",
        )

    status_filter = status if status else None

    esg_task = db.get_esg_summary(
        country=country, business_size=business_size, status=status_filter,
        columns="status,is_small_scale_farmer,key_person_count,female_director_count,youth_director_count",
    )
    size_task = db.get_business_size_distribution()
    cat_task = db.get_category_compliance_stats(country=country, business_size=business_size, status=status_filter)
    doc_task = db.get_document_type_stats()
    supplier_task = db.get_sustainability_supplier_list(country=country, business_size=business_size, status=status_filter)

    esg_rows, size_rows, cat_rows, doc_rows, supplier_rows = await asyncio.gather(
        esg_task, size_task, cat_task, doc_task, supplier_task
    )

    if status_filter is None:
        esg_rows = [r for r in esg_rows if (r.get("status") or "") in APPROVED_STATUS_SCOPE]

    total = len(esg_rows)
    status_breakdown = {
        "approved": sum(1 for r in esg_rows if (r.get("status") or "") == "APPROVED"),
        "compliance_required": sum(1 for r in esg_rows if (r.get("status") or "") == "COMPLIANCE_REQUIRED"),
        "suspended": sum(1 for r in esg_rows if (r.get("status") or "") == "SUSPENDED"),
    }
    derived_flags = [_derive_esg_booleans_from_counts(r) for r in esg_rows]
    women_owned = sum(1 for wo, _ in derived_flags if wo)
    youth_owned = sum(1 for _, yo in derived_flags if yo)
    farmers = sum(1 for r in esg_rows if r.get("is_small_scale_farmer"))
    female_directors = sum(r.get("female_director_count") or 0 for r in esg_rows)
    youth_directors = sum(r.get("youth_director_count") or 0 for r in esg_rows)
    total_directors = sum(r.get("key_person_count") or 0 for r in esg_rows)

    now_str = get_cat_now().strftime("%B %d, %Y at %I:%M %p CAT")

    # ── Shared styles ────────────────────────────────────────────────────────
    def _thin_border() -> Border:
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr_cell(ws, row: int, col: int, value: str, fill_hex: str) -> None:
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()

    def _data_cell(ws, row: int, col: int, value, alt: bool = False, fill_hex: str = "") -> None:
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = _thin_border()
        if fill_hex:
            c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        elif alt:
            c.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    wb = openpyxl.Workbook()

    # ══ Sheet 1: ESG KPIs ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "ESG KPIs"
    ws1.row_dimensions[1].height = 28
    ws1.freeze_panes = "A4"

    # Report header
    ws1["A1"] = "RTG Procurement — Sustainability & ESG Report"
    ws1["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws1["A2"] = f"Generated: {now_str}"
    ws1["A2"].font = Font(italic=True, size=10, color="6B7280")
    filters_applied = []
    if country:
        filters_applied.append(f"Country: {country}")
    if business_size:
        filters_applied.append(f"Business Size: {business_size}")
    if status:
        filters_applied.append(f"Status: {status}")
    ws1["A3"] = "Filters: " + (", ".join(filters_applied) if filters_applied else "All Approved Suppliers")
    ws1["A3"].font = Font(italic=True, size=10, color="374151")

    # Status breakdown mini-table
    kpi_section_start = 5
    ws1.cell(row=kpi_section_start, column=1, value="Status Breakdown").font = Font(bold=True, size=12, color="1E3A5F")
    status_hdrs = ["Total in Scope", "Approved", "Compliance Required", "Suspended"]
    status_vals = [total, status_breakdown["approved"], status_breakdown["compliance_required"], status_breakdown["suspended"]]
    status_fills = ["1E3A5F", "059669", "D97706", "DC2626"]
    for ci, (hdr, val, fx) in enumerate(zip(status_hdrs, status_vals, status_fills), 1):
        ws1.cell(row=kpi_section_start + 1, column=ci, value=hdr).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=kpi_section_start + 1, column=ci).fill = PatternFill(start_color=fx, end_color=fx, fill_type="solid")
        ws1.cell(row=kpi_section_start + 1, column=ci).alignment = Alignment(horizontal="center")
        ws1.cell(row=kpi_section_start + 1, column=ci).border = _thin_border()
        c = ws1.cell(row=kpi_section_start + 2, column=ci, value=val)
        c.font = Font(bold=True, size=14, color=fx)
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()

    # ESG metrics table
    esg_section_start = kpi_section_start + 5
    ws1.cell(row=esg_section_start, column=1, value="ESG Metrics").font = Font(bold=True, size=12, color="1E3A5F")
    esg_hdrs = ["ESG Metric", "Count", "% of Suppliers"]
    esg_data = [
        ("Women-Owned / Led (>50% female key persons)", women_owned, f"{_safe_pct(women_owned, total)}%"),
        ("Youth-Owned / Led (>50% youth key persons, aged <35)", youth_owned, f"{_safe_pct(youth_owned, total)}%"),
        ("Small-Scale Farmers", farmers, f"{_safe_pct(farmers, total)}%"),
        ("Female Directors", female_directors, f"{_safe_pct(female_directors, total_directors)}% of directors"),
        ("Youth Directors (aged <35)", youth_directors, f"{_safe_pct(youth_directors, total_directors)}% of directors"),
        ("Total Directors / Key Persons", total_directors, "—"),
    ]
    esg_fills = ["EC4899", "7C3AED", "059669", "F59E0B", "6366F1", "6B7280"]
    for ci, hdr in enumerate(esg_hdrs, 1):
        _hdr_cell(ws1, esg_section_start + 1, ci, hdr, "0F766E")
    for ri, ((label, val, pct), fx) in enumerate(zip(esg_data, esg_fills), start=esg_section_start + 2):
        ws1.cell(row=ri, column=1, value=label).alignment = Alignment(vertical="center", wrap_text=True)
        ws1.cell(row=ri, column=1).border = _thin_border()
        c_val = ws1.cell(row=ri, column=2, value=val)
        c_val.font = Font(bold=True, size=12, color=fx)
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = _thin_border()
        ws1.cell(row=ri, column=3, value=pct).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=ri, column=3).border = _thin_border()
        if ri % 2 == 0:
            for ci in range(1, 4):
                ws1.cell(row=ri, column=ci).fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    ws1.column_dimensions["A"].width = 52
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 22

    # ══ Sheet 2: Business Size ════════════════════════════════════════════════
    ws2 = wb.create_sheet("Business Size")
    ws2.freeze_panes = "A3"
    ws2["A1"] = "Business Size Distribution"
    ws2["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws2["A2"] = f"Generated: {now_str}"
    ws2["A2"].font = Font(italic=True, size=10, color="6B7280")

    size_hdrs = ["Business Size", "Definition", "Supplier Count", "% of Total"]
    size_defs = {"SMALL": "< 10 employees", "MEDIUM": "10 – 50 employees", "LARGE": "> 50 employees"}
    size_fills_map = {"SMALL": "D1FAE5", "MEDIUM": "EDE9FE", "LARGE": "FEF3C7"}
    for ci, hdr in enumerate(size_hdrs, 1):
        _hdr_cell(ws2, 4, ci, hdr, "2563EB")

    size_total = sum(r.get("supplier_count", 0) for r in size_rows)
    small_c = medium_c = large_c = 0
    for ri, r in enumerate(sorted(size_rows, key=lambda x: ["SMALL", "MEDIUM", "LARGE", "UNKNOWN"].index(x.get("business_size", "UNKNOWN")) if x.get("business_size", "UNKNOWN") in ["SMALL", "MEDIUM", "LARGE"] else 3), start=5):
        sz = r.get("business_size") or "UNKNOWN"
        cnt = r.get("supplier_count", 0)
        if sz == "SMALL": small_c = cnt
        elif sz == "MEDIUM": medium_c = cnt
        elif sz == "LARGE": large_c = cnt
        row_fill = size_fills_map.get(sz, "F9FAFB")
        ws2.cell(row=ri, column=1, value=sz.title()).fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        ws2.cell(row=ri, column=1).border = _thin_border()
        ws2.cell(row=ri, column=1).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=size_defs.get(sz, "No data")).border = _thin_border()
        ws2.cell(row=ri, column=2).fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        c = ws2.cell(row=ri, column=3, value=cnt)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()
        c.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        ws2.cell(row=ri, column=4, value=f"{_safe_pct(cnt, size_total)}%").alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=4).border = _thin_border()
        ws2.cell(row=ri, column=4).fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")

    # SME summary row
    sme_total = small_c + medium_c
    sme_row = 5 + len(size_rows)
    ws2.cell(row=sme_row + 1, column=1, value="Total SMEs (Small + Medium)").font = Font(bold=True, size=11, color="1D4ED8")
    ws2.cell(row=sme_row + 1, column=1).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws2.cell(row=sme_row + 1, column=1).border = _thin_border()
    ws2.cell(row=sme_row + 1, column=2, value="<= 50 employees").border = _thin_border()
    ws2.cell(row=sme_row + 1, column=2).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws2.cell(row=sme_row + 1, column=3, value=sme_total).font = Font(bold=True, size=13, color="1D4ED8")
    ws2.cell(row=sme_row + 1, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=sme_row + 1, column=3).border = _thin_border()
    ws2.cell(row=sme_row + 1, column=3).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws2.cell(row=sme_row + 1, column=4, value=f"{_safe_pct(sme_total, size_total)}%").alignment = Alignment(horizontal="center")
    ws2.cell(row=sme_row + 1, column=4).border = _thin_border()
    ws2.cell(row=sme_row + 1, column=4).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for ci, w in enumerate([28, 22, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ══ Sheet 3: Category Compliance ═════════════════════════════════════════
    ws3 = wb.create_sheet("Category Compliance")
    ws3.freeze_panes = "A4"
    ws3["A1"] = "Category Compliance Overview"
    ws3["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws3["A2"] = f"Generated: {now_str}  |  Sorted by compliance % descending"
    ws3["A2"].font = Font(italic=True, size=10, color="6B7280")

    cc_hdrs = ["Business Category", "Total Suppliers", "Mand. Met", "Mand. Missing",
               "Full Compliance", "Med Risk", "High Risk", "Pending", "Excluded", "Compliance %"]
    cc_fill = "1D4ED8"
    for ci, hdr in enumerate(cc_hdrs, 1):
        _hdr_cell(ws3, 3, ci, hdr, cc_fill)

    sorted_cats = sorted(
        [r for r in cat_rows if r.get("total_suppliers", 0) > 0],
        key=lambda r: r.get("full_compliance_pct") or 0, reverse=True
    )
    for ri, r in enumerate(sorted_cats, start=4):
        pct = r.get("full_compliance_pct") or 0
        alt = ri % 2 == 0
        alt_fill = "EFF6FF" if alt else ""
        ws3.cell(row=ri, column=1, value=r["category"].replace("_", " ").title())
        ws3.cell(row=ri, column=1).border = _thin_border()
        ws3.cell(row=ri, column=1).font = Font(bold=True)
        if alt_fill:
            ws3.cell(row=ri, column=1).fill = PatternFill(start_color=alt_fill, end_color=alt_fill, fill_type="solid")

        fields = [
            r["total_suppliers"], r.get("mandatory_met_count", 0), r.get("mandatory_missing_count", 0),
            r["full_compliance_count"], r["medium_risk_count"], r["high_risk_count"],
            r["pending_count"], r.get("excluded_count", 0),
        ]
        for ci, val in enumerate(fields, 2):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
            if alt_fill:
                c.fill = PatternFill(start_color=alt_fill, end_color=alt_fill, fill_type="solid")
            # Highlight problems
            if ci == 4 and val > 0:  # mandatory missing
                c.font = Font(bold=True, color="DC2626")
            elif ci == 7 and val > 0:  # high risk
                c.font = Font(color="DC2626")
            elif ci == 6 and val > 0:  # medium risk
                c.font = Font(color="D97706")

        # Compliance % with color
        pct_str = f"{pct:.1f}%"
        c_pct = ws3.cell(row=ri, column=10, value=pct_str)
        c_pct.alignment = Alignment(horizontal="center", vertical="center")
        c_pct.border = _thin_border()
        if pct >= 75:
            c_pct.font = Font(bold=True, color="059669")
            c_pct.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif pct >= 40:
            c_pct.font = Font(bold=True, color="D97706")
            c_pct.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            c_pct.font = Font(bold=True, color="DC2626")
            c_pct.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    cc_col_widths = [36, 14, 12, 14, 14, 10, 10, 10, 10, 14]
    for ci, w in enumerate(cc_col_widths, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ══ Sheet 4: Document Stats ═══════════════════════════════════════════════
    ws4 = wb.create_sheet("Document Stats")
    ws4.freeze_panes = "A4"
    ws4["A1"] = "Document Coverage & Verification Status"
    ws4["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws4["A2"] = f"Generated: {now_str}  |  Sorted by supplier count descending"
    ws4["A2"].font = Font(italic=True, size=10, color="6B7280")

    ds_hdrs = ["Document Type", "Suppliers", "Total Uploads", "Verified", "Pending", "Rejected", "Verification Rate"]
    for ci, hdr in enumerate(ds_hdrs, 1):
        _hdr_cell(ws4, 3, ci, hdr, "0F766E")

    for ri, r in enumerate(sorted(doc_rows, key=lambda x: x.get("supplier_count", 0), reverse=True), start=4):
        alt = ri % 2 == 0
        total_ups = r.get("total_uploads") or 0
        verified = r.get("verified_count") or 0
        ver_rate = _safe_pct(verified, total_ups) if total_ups else 0
        row_data = [
            (r.get("document_type") or "").replace("_", " ").title(),
            r.get("supplier_count", 0),
            total_ups,
            verified,
            r.get("pending_count", 0),
            r.get("rejected_count", 0),
            f"{ver_rate}%" if total_ups else "—",
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center", wrap_text=(ci == 1))
            if alt:
                c.fill = PatternFill(start_color="ECFEFF", end_color="ECFEFF", fill_type="solid")
            # Color verification rate
            if ci == 7 and total_ups > 0:
                if ver_rate >= 75:
                    c.font = Font(bold=True, color="059669")
                elif ver_rate >= 40:
                    c.font = Font(bold=True, color="D97706")
                else:
                    c.font = Font(bold=True, color="DC2626")
            # Red pending/rejected if high
            elif ci == 5 and isinstance(val, int) and val > 0:
                c.font = Font(color="D97706")
            elif ci == 6 and isinstance(val, int) and val > 0:
                c.font = Font(color="DC2626")

    ds_col_widths = [40, 12, 14, 12, 12, 12, 16]
    for ci, w in enumerate(ds_col_widths, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ══ Sheet 5: Supplier List ════════════════════════════════════════════════
    ws5 = wb.create_sheet("Supplier List")
    ws5.freeze_panes = "A3"
    ws5["A1"] = f"Supplier List — {len(supplier_rows)} supplier(s)"
    ws5["A1"].font = Font(bold=True, size=13, color="1E3A5F")

    sl_cols = [
        ("Company Name", "company_name"),
        ("Country", "country"),
        ("Status", "status"),
        ("Business Size", "business_size"),
        ("Employees", "employee_count"),
        ("Women-Owned", "esg_women_owned"),
        ("Youth-Owned", "esg_youth_owned"),
        ("Small-Scale Farmer", "is_small_scale_farmer"),
        ("Categories", "business_categories"),
        ("Years in Business", "years_in_business"),
        ("Registered", "created_at"),
    ]
    status_color_map = {
        "APPROVED": "059669", "COMPLIANCE_REQUIRED": "D97706",
        "SUSPENDED": "DC2626", "REJECTED": "DC2626",
    }
    for ci, (hdr, _) in enumerate(sl_cols, 1):
        _hdr_cell(ws5, 2, ci, hdr, "374151")

    for ri, row in enumerate(supplier_rows, start=3):
        alt = ri % 2 == 0
        supplier_status = (row.get("status") or "").upper()
        for ci, (_, key) in enumerate(sl_cols, 1):
            val = row.get(key)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val).replace("_", " ")
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            c = ws5.cell(row=ri, column=ci, value=val)
            c.border = _thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=(ci in (1, 9)))
            if alt:
                c.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            if ci == 3 and supplier_status in status_color_map:
                c.font = Font(bold=True, color=status_color_map[supplier_status])
            if ci in (6, 7, 8):
                c.alignment = Alignment(horizontal="center", vertical="center")

    sl_col_widths = [32, 16, 18, 14, 11, 13, 12, 16, 40, 14, 18]
    for ci, w in enumerate(sl_col_widths, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"rtg_sustainability_report_{get_cat_timestamp_str()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 5. PDF download (comprehensive full report) ───────────────────────────────

@router.get("/export/pdf", summary="Download comprehensive sustainability PDF report")
async def export_sustainability_pdf(
    country: Optional[str] = Query(None),
    business_size: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_admin: dict = Depends(get_current_admin),
):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable,
        )
    except ImportError:
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="reportlab is required for PDF export. Run: pip install reportlab",
        )

    status_filter = status if status else None

    esg_task = db.get_esg_summary(
        country=country, business_size=business_size, status=status_filter,
        columns="status,is_small_scale_farmer,key_person_count,female_director_count,youth_director_count",
    )
    size_task = db.get_business_size_distribution()
    cat_task = db.get_category_compliance_stats(country=country, business_size=business_size, status=status_filter)
    doc_task = db.get_document_type_stats()
    supplier_task = db.get_sustainability_supplier_list(country=country, business_size=business_size, status=status_filter)

    esg_rows, size_rows, cat_rows, doc_rows, supplier_rows = await asyncio.gather(
        esg_task, size_task, cat_task, doc_task, supplier_task
    )

    if status_filter is None:
        esg_rows = [r for r in esg_rows if (r.get("status") or "") in APPROVED_STATUS_SCOPE]

    total = len(esg_rows)
    status_breakdown = {
        "approved": sum(1 for r in esg_rows if (r.get("status") or "") == "APPROVED"),
        "compliance_required": sum(1 for r in esg_rows if (r.get("status") or "") == "COMPLIANCE_REQUIRED"),
        "suspended": sum(1 for r in esg_rows if (r.get("status") or "") == "SUSPENDED"),
    }
    derived_flags = [_derive_esg_booleans_from_counts(r) for r in esg_rows]
    women_owned = sum(1 for wo, _ in derived_flags if wo)
    youth_owned = sum(1 for _, yo in derived_flags if yo)
    farmers = sum(1 for r in esg_rows if r.get("is_small_scale_farmer"))
    female_directors = sum(r.get("female_director_count") or 0 for r in esg_rows)
    youth_directors = sum(r.get("youth_director_count") or 0 for r in esg_rows)
    total_directors = sum(r.get("key_person_count") or 0 for r in esg_rows)

    now_str = get_cat_now().strftime("%B %d, %Y at %I:%M %p CAT")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=0.6 * inch, leftMargin=0.6 * inch,
        topMargin=0.7 * inch, bottomMargin=0.6 * inch,
        title="RTG Sustainability & ESG Report",
        author="RTG Procurement Portal",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RTGTitle", parent=styles["Normal"],
        fontSize=20, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0F172A"), spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "RTGSub", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        textColor=colors.HexColor("#6B7280"), spaceAfter=2,
    )
    h2_style = ParagraphStyle(
        "RTGH2", parent=styles["Normal"],
        fontSize=12, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E3A5F"), spaceBefore=14, spaceAfter=6,
    )
    normal_style = ParagraphStyle(
        "RTGNormal", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica",
        textColor=colors.HexColor("#374151"),
    )

    PAGE_W = landscape(A4)[0] - 1.2 * inch  # usable width

    elements = []

    # ── Cover ─────────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("RTG Procurement — Sustainability &amp; ESG Report", title_style))
    elements.append(Paragraph(f"Generated: {now_str}", sub_style))
    filters_applied_str = ", ".join(
        [f"Country: {country}"] * bool(country)
        + [f"Business Size: {business_size}"] * bool(business_size)
        + [f"Status: {status}"] * bool(status)
    ) or "All Approved Suppliers"
    elements.append(Paragraph(f"Filters: {filters_applied_str}", sub_style))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1E3A5F")))
    elements.append(Spacer(1, 0.15 * inch))

    # Headline KPI summary table
    kpi_hdrs = ["Total in Scope", "Approved", "Compliance Req.", "Suspended",
                "Women-Owned/Led", "Youth-Owned/Led", "Small-Scale Farmers"]
    kpi_vals = [
        str(total),
        str(status_breakdown["approved"]),
        str(status_breakdown["compliance_required"]),
        str(status_breakdown["suspended"]),
        f"{women_owned}\n({_safe_pct(women_owned, total)}%)",
        f"{youth_owned}\n({_safe_pct(youth_owned, total)}%)",
        f"{farmers}\n({_safe_pct(farmers, total)}%)",
    ]
    kpi_tbl = Table([kpi_hdrs, kpi_vals], colWidths=[PAGE_W / 7] * 7)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 14),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#059669")),
        ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#D97706")),
        ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#DC2626")),
        ("TEXTCOLOR", (4, 1), (4, 1), colors.HexColor("#EC4899")),
        ("TEXTCOLOR", (5, 1), (5, 1), colors.HexColor("#7C3AED")),
        ("TEXTCOLOR", (6, 1), (6, 1), colors.HexColor("#059669")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("BOX", (0, 0), (-1, -1), 1.2, colors.HexColor("#1E3A5F")),
    ]))
    elements.append(kpi_tbl)
    elements.append(PageBreak())

    # ── 1. ESG Metrics ────────────────────────────────────────────────────────
    elements.append(Paragraph("1. ESG Metrics", h2_style))
    esg_tbl_data = [
        ["ESG Metric", "Count", "Percentage"],
        ["Women-Owned / Led Suppliers (>50% female key persons)", str(women_owned), f"{_safe_pct(women_owned, total)}%"],
        ["Youth-Owned / Led Suppliers (>50% youth key persons, aged <35)", str(youth_owned), f"{_safe_pct(youth_owned, total)}%"],
        ["Small-Scale Farmers", str(farmers), f"{_safe_pct(farmers, total)}%"],
        ["Female Directors", str(female_directors), f"{_safe_pct(female_directors, total_directors)}% of directors"],
        ["Youth Directors (aged <35)", str(youth_directors), f"{_safe_pct(youth_directors, total_directors)}% of directors"],
        ["Total Directors / Key Persons", str(total_directors), "—"],
    ]
    esg_tbl = Table(esg_tbl_data, colWidths=[PAGE_W * 0.6, PAGE_W * 0.15, PAGE_W * 0.25], repeatRows=1)
    esg_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ecfeff")]),
    ]))
    elements.append(esg_tbl)

    # ── 2. Business Size ──────────────────────────────────────────────────────
    elements.append(Paragraph("2. Business Size Distribution", h2_style))
    small_c = next((r["supplier_count"] for r in size_rows if r.get("business_size") == "SMALL"), 0)
    medium_c = next((r["supplier_count"] for r in size_rows if r.get("business_size") == "MEDIUM"), 0)
    large_c = next((r["supplier_count"] for r in size_rows if r.get("business_size") == "LARGE"), 0)
    size_total = max(sum(r.get("supplier_count", 0) for r in size_rows), 1)
    size_tbl_data = [
        ["Business Size", "Definition", "Supplier Count", "% of Total"],
        ["Small Enterprise", "< 10 employees", str(small_c), f"{_safe_pct(small_c, size_total)}%"],
        ["Medium Enterprise", "10 – 50 employees", str(medium_c), f"{_safe_pct(medium_c, size_total)}%"],
        ["Large Enterprise", "> 50 employees", str(large_c), f"{_safe_pct(large_c, size_total)}%"],
        ["Total SMEs (Small + Medium)", "<= 50 employees", str(small_c + medium_c), f"{_safe_pct(small_c + medium_c, size_total)}%"],
    ]
    size_tbl = Table(size_tbl_data, colWidths=[PAGE_W * 0.3, PAGE_W * 0.25, PAGE_W * 0.22, PAGE_W * 0.23], repeatRows=1)
    size_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (1, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#eff6ff")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 9),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#1D4ED8")),
    ]))
    elements.append(size_tbl)
    elements.append(PageBreak())

    # ── 3. Category Compliance ────────────────────────────────────────────────
    elements.append(Paragraph("3. Category Compliance Overview", h2_style))
    elements.append(Paragraph(
        "Sorted by full compliance % descending. Red = non-compliant. Amber = medium risk. Green = 75%+ compliance.",
        normal_style,
    ))
    elements.append(Spacer(1, 0.08 * inch))

    sorted_cats = sorted(
        [r for r in cat_rows if r.get("total_suppliers", 0) > 0],
        key=lambda r: r.get("full_compliance_pct") or 0, reverse=True
    )
    cc_data = [["Business Category", "Total", "Mand. Met", "Mand. Missing",
                "Full Compliance", "Med Risk", "High Risk", "Pending", "Compliance %"]]
    for r in sorted_cats:
        cc_data.append([
            r["category"].replace("_", " ").title()[:32],
            str(r["total_suppliers"]),
            str(r.get("mandatory_met_count", 0)),
            str(r.get("mandatory_missing_count", 0)),
            str(r["full_compliance_count"]),
            str(r["medium_risk_count"]),
            str(r["high_risk_count"]),
            str(r["pending_count"]),
            f"{r['full_compliance_pct']:.1f}%",
        ])
    if len(cc_data) > 1:
        cc_col_ws = [PAGE_W * 0.24, PAGE_W * 0.06, PAGE_W * 0.08, PAGE_W * 0.1,
                     PAGE_W * 0.09, PAGE_W * 0.08, PAGE_W * 0.08, PAGE_W * 0.07, PAGE_W * 0.1]
        cc_tbl = Table(cc_data, colWidths=cc_col_ws, repeatRows=1)
        cc_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        ]
        for ri, r in enumerate(sorted_cats, start=1):
            pct = r.get("full_compliance_pct") or 0
            if pct >= 75:
                cc_style.append(("TEXTCOLOR", (8, ri), (8, ri), colors.HexColor("#059669")))
                cc_style.append(("FONTNAME", (8, ri), (8, ri), "Helvetica-Bold"))
            elif pct >= 40:
                cc_style.append(("TEXTCOLOR", (8, ri), (8, ri), colors.HexColor("#D97706")))
                cc_style.append(("FONTNAME", (8, ri), (8, ri), "Helvetica-Bold"))
            else:
                cc_style.append(("TEXTCOLOR", (8, ri), (8, ri), colors.HexColor("#DC2626")))
                cc_style.append(("FONTNAME", (8, ri), (8, ri), "Helvetica-Bold"))
            if r.get("high_risk_count", 0) > 0:
                cc_style.append(("TEXTCOLOR", (6, ri), (6, ri), colors.HexColor("#DC2626")))
            if r.get("mandatory_missing_count", 0) > 0:
                cc_style.append(("TEXTCOLOR", (3, ri), (3, ri), colors.HexColor("#DC2626")))
                cc_style.append(("FONTNAME", (3, ri), (3, ri), "Helvetica-Bold"))
        cc_tbl.setStyle(TableStyle(cc_style))
        elements.append(cc_tbl)
    else:
        elements.append(Paragraph("No category compliance data available.", normal_style))
    elements.append(PageBreak())

    # ── 4. Document Statistics ────────────────────────────────────────────────
    elements.append(Paragraph("4. Document Coverage &amp; Verification Status", h2_style))
    elements.append(Paragraph(
        "Counts distinct suppliers with uploaded documents per type (non-archived only). Sorted by supplier count descending.",
        normal_style,
    ))
    elements.append(Spacer(1, 0.08 * inch))

    ds_data = [["Document Type", "Suppliers", "Total Uploads", "Verified", "Pending", "Rejected", "Verification Rate"]]
    for r in sorted(doc_rows, key=lambda x: x.get("supplier_count", 0), reverse=True):
        total_ups = r.get("total_uploads") or 0
        verified = r.get("verified_count") or 0
        ds_data.append([
            (r.get("document_type") or "").replace("_", " ").title()[:38],
            str(r.get("supplier_count", 0)),
            str(total_ups),
            str(verified),
            str(r.get("pending_count", 0)),
            str(r.get("rejected_count", 0)),
            f"{_safe_pct(verified, total_ups)}%" if total_ups else "—",
        ])
    if len(ds_data) > 1:
        ds_col_ws = [PAGE_W * 0.32, PAGE_W * 0.1, PAGE_W * 0.11, PAGE_W * 0.1,
                     PAGE_W * 0.1, PAGE_W * 0.1, PAGE_W * 0.17]
        ds_tbl = Table(ds_data, colWidths=ds_col_ws, repeatRows=1)
        ds_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECFEFF")]),
        ]
        for ri, r in enumerate(sorted(doc_rows, key=lambda x: x.get("supplier_count", 0), reverse=True), start=1):
            total_ups = r.get("total_uploads") or 0
            verified = r.get("verified_count") or 0
            ver_rate = _safe_pct(verified, total_ups) if total_ups else 0
            if total_ups > 0:
                if ver_rate >= 75:
                    ds_style.append(("TEXTCOLOR", (6, ri), (6, ri), colors.HexColor("#059669")))
                elif ver_rate >= 40:
                    ds_style.append(("TEXTCOLOR", (6, ri), (6, ri), colors.HexColor("#D97706")))
                else:
                    ds_style.append(("TEXTCOLOR", (6, ri), (6, ri), colors.HexColor("#DC2626")))
        ds_tbl.setStyle(TableStyle(ds_style))
        elements.append(ds_tbl)
    else:
        elements.append(Paragraph("No document statistics available.", normal_style))
    elements.append(PageBreak())

    # ── 5. Supplier List ──────────────────────────────────────────────────────
    if supplier_rows:
        elements.append(Paragraph(f"5. Supplier List ({len(supplier_rows)} suppliers)", h2_style))
        sl_data = [["Company", "Country", "Status", "Size", "Women", "Youth", "Farmer", "Categories"]]
        for s in supplier_rows:
            cats = s.get("business_categories") or []
            cats_str = (", ".join(cats) if isinstance(cats, list) else str(cats)).replace("_", " ")[:45]
            sl_data.append([
                (s.get("company_name") or "")[:30],
                (s.get("country") or "")[:14],
                (s.get("status") or "")[:14],
                (s.get("business_size") or "—")[:8],
                "Yes" if s.get("esg_women_owned") else "No",
                "Yes" if s.get("esg_youth_owned") else "No",
                "Yes" if s.get("is_small_scale_farmer") else "No",
                cats_str,
            ])
        sl_col_ws = [PAGE_W * 0.22, PAGE_W * 0.1, PAGE_W * 0.1, PAGE_W * 0.07,
                     PAGE_W * 0.06, PAGE_W * 0.06, PAGE_W * 0.07, PAGE_W * 0.32]
        sl_tbl = Table(sl_data, colWidths=sl_col_ws, repeatRows=1)
        sl_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (2, 0), (6, -1), "CENTER"),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("ALIGN", (7, 0), (7, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]
        status_color_pdf = {
            "APPROVED": colors.HexColor("#059669"),
            "COMPLIANCE_REQUIRED": colors.HexColor("#D97706"),
            "SUSPENDED": colors.HexColor("#DC2626"),
            "REJECTED": colors.HexColor("#DC2626"),
        }
        for ri, s in enumerate(supplier_rows, start=1):
            st = (s.get("status") or "").upper()
            if st in status_color_pdf:
                sl_style.append(("TEXTCOLOR", (2, ri), (2, ri), status_color_pdf[st]))
                sl_style.append(("FONTNAME", (2, ri), (2, ri), "Helvetica-Bold"))
        sl_tbl.setStyle(TableStyle(sl_style))
        elements.append(sl_tbl)

    doc.build(elements)
    buf.seek(0)

    filename = f"rtg_sustainability_report_{get_cat_timestamp_str()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
